
import openpyxl
from os import system
from io import BytesIO
from requests import Session
from bs4 import BeautifulSoup
from PIL import Image as PILImage
from openpyxl.drawing.image import Image


class Amazon:
    def __init__(self):
        self.net = Session()
        self.condicao = "s?i=fashion-boys-intl-ship&bbn"
        self.url_root = "https://www.amazon.com/"
        self.net.headers["Accept"] = "text/htmltext/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
        self.net.headers["Accept-Language"] = "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7"
        self.net.headers["User-Agent"] = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36"

    def iniciar(self):
        global numero
        system("mkdir Imagens")
        site = self.net.get(self.url_root+self.condicao+"=16225021011&rh=n%3A16225021011%2Cn%3A1040666&dc&page=1&language=pt_BR&fst")
        soup = BeautifulSoup(site.text, 'html.parser')
        # Motra o número de páginas 
        paginas = soup.find('span', class_='s-pagination-item s-pagination-disabled').text
        numero = paginas

        print("=="*30)
        print(f"{f'O Site tem {paginas} paginas'.center(60)}")
        print("=="*30)
        dados = self.varrer_site()
        self.adicionar_dados_a_planilha(dados)

    def adicionar_dados_a_planilha(self, dados):
	    # Criar a planilha
	    workbook = openpyxl.Workbook()
	    sheet = workbook.active

	    # Adicionar o cabeçalho da planilha
	    sheet['A1'] = 'Nome'
	    sheet['B1'] = 'Preço'
	    sheet['C1'] = 'Aumento'
	    sheet['D1'] = 'Imagem'

	    # Adicionar os dados à planilha
	    for i, produto in enumerate(dados):
	        # Baixar a imagem a partir da URL e abrir com o Pillow
	        response = self.net.get(produto['imagem'])
	        img = PILImage.open(BytesIO(response.content))

	        # Adicionar o nome e o preço na planilha
	        sheet.cell(row=i+2, column=1).value = produto['nome']
	        sheet.cell(row=i+2, column=2).value = produto['preco']
	        sheet.cell(row=i+2, column=3).value = produto['aumento']

	        # Adicionar a imagem à planilha
	        img_path = 'Imagens/imagem_{}.jpeg'.format(i)
	        img.save(img_path)
	        img = Image(img_path)
	        sheet.column_dimensions['D'].width = 30
	        sheet.row_dimensions[i+2].height = 80
	        sheet.add_image(img, 'D{}'.format(i+2))

	    # Salvar a planilha
	    workbook.save('dados.xlsx')  
    
    @staticmethod
    def aumentar_percentual(valor, percentual):
        aumento = valor * (percentual / 100)
        valor_final = valor + aumento
        return valor_final

    def varrer_site(self):
        c = 1
        lista = []
        global numero
        dicionario = {}
        
        
        while True:
            site = self.net.get(f"{self.url_root}{self.condicao}=16225021011&rh=n%3A16225021011%2Cn%3A1040666&dc&page={c}&language=pt_BR&fst")
            soup = BeautifulSoup(site.text, 'html.parser')
            blocos = soup.find_all('div', class_='a-section a-spacing-base')
            for posicao in range(0, len(blocos)):
                nome = blocos[posicao].find('div', class_='a-section a-spacing-none a-spacing-top-small s-title-instructions-style').text
                # print(nome)
                try:
                    simbolo = blocos[posicao].find("span", class_='a-price-symbol').text
                    valor = blocos[posicao].find("span", class_='a-price-whole').text
                    decimal = blocos[posicao].find("span", class_='a-price-fraction').text
                    preco = f"{simbolo}{valor}{decimal}"
                    url_imagem = blocos[posicao].find("img", class_='s-image').get("src")
                except:
                    pass
                
                aumento = self.aumentar_percentual(int(str(valor).replace(",", "")), 20)
                dicionario = {"nome": nome, "preco": preco, "aumento": f"{simbolo}{aumento}{decimal}", "imagem": url_imagem}
                lista.append(dicionario)

            print(f"Varrendo a página {c}")
            
            c += 1
            if c >= int(numero):
                break
        return lista
        
numero = 0
system("clear || cls")
amazon = Amazon()
amazon.iniciar()
